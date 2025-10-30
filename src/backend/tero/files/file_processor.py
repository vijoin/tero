from abc import ABC, abstractmethod
import io
from io import BytesIO
import logging
from typing import Optional, Any

import openpyxl
import openpyxl.worksheet.worksheet
from PIL import Image
import xlrd

from ..files.domain import File
from ..files.file_quota import FileQuota
from ..files.pdf_processor import process_pdf_basic, process_pdf_enhanced


logger = logging.getLogger(__name__)

def get_encoding(content_type: Optional[str]) -> str:
    charset_param = '; charset='
    encoding = content_type.split(charset_param, 1)[1] if content_type and charset_param in content_type else 'utf-8'
    return encoding

class BaseFileProcessor(ABC):    
    @abstractmethod
    def supports(self, file: File) -> bool:
        # Checks if this processor supports the given file
        pass
    
    @abstractmethod
    def extract_text(self, file: File, file_quota: FileQuota) -> str:
        pass

class PlainTextFileProcessor(BaseFileProcessor):

    def supports(self, file: File) -> bool:
        return any(file.name.lower().endswith(ext) for ext in {'.txt', '.md', '.csv', '.har', '.json', '.svg'})
    
    def extract_text(self, file: File, file_quota: FileQuota) -> str:
        encoding = get_encoding(file.content_type)
        return file.content.decode(encoding)

class Sheet(ABC):
    
    @property
    @abstractmethod
    def title(self) -> str:
        pass

    @property
    @abstractmethod
    def row_count(self) -> int:
        pass

    @property
    @abstractmethod
    def column_count(self) -> int:
        pass

    @abstractmethod
    def cell(self, row_idx: int, col_idx: int) -> Any:
        pass

class SpreadsheetFileProcessor(BaseFileProcessor, ABC):
    file_extension: str

    def supports(self, file: File) -> bool:
        return file.name.lower().endswith(self.file_extension)

    def extract_text(self, file: File, file_quota: FileQuota) -> str:
        sheets = self._load_sheets(file.content)
        sheet_count = sum(1 for sheet in sheets if sheet.row_count)
        return "\n\n".join(self._format_sheet(sheet, include_sheet_name=sheet_count > 1) for sheet in sheets if sheet.row_count)

    @abstractmethod
    def _load_sheets(self, content: bytes) -> list[Sheet]:
        pass

    def _format_sheet(self, sheet: Sheet, include_sheet_name: bool) -> str:
        prefix = f"## Sheet {sheet.title}\n\n" if include_sheet_name else ""
        return prefix + "\n".join(self._format_row(row_idx, sheet) for row_idx in range(sheet.row_count))

    def _format_row(self, row_idx: int, sheet: Sheet) -> str:
        return " | ".join(self._format_cell(row_idx, col_idx, sheet) for col_idx in range(sheet.column_count))

    def _format_cell(self, row_idx: int, col_idx: int, sheet: Sheet) -> str:
        ret = sheet.cell(row_idx, col_idx)
        return str(ret) if ret is not None else ""

class XlsxSheet(Sheet):

    def __init__(self, sheet:openpyxl.worksheet.worksheet.Worksheet):
        self._sheet = sheet

    @property
    def title(self) -> str:
        return self._sheet.title
    
    @property
    def row_count(self) -> int:
        return self._sheet.max_row
    
    @property
    def column_count(self) -> int:
        return self._sheet.max_column
    
    def cell(self, row_idx: int, col_idx: int) -> Any:
        return self._sheet.cell(row_idx + 1, col_idx + 1).value

class XlsxFileProcessor(SpreadsheetFileProcessor):
    file_extension = '.xlsx'

    def _load_sheets(self, content: bytes) -> list[Sheet]:
        wb = openpyxl.load_workbook(BytesIO(content))
        return [XlsxSheet(sheet) for sheet in wb.worksheets]

class XlsSheet(Sheet):

    def __init__(self, sheet: xlrd.sheet.Sheet):
        self._sheet = sheet

    def is_empyt(self) -> bool:
        return self._sheet.nrows == 0

    @property
    def title(self) -> str:
        return self._sheet.name

    @property
    def row_count(self) -> int:
        return self._sheet.nrows

    @property
    def column_count(self) -> int:
        return self._sheet.ncols

    def cell(self, row_idx: int, col_idx: int) -> Any:
        return self._sheet.cell(row_idx, col_idx).value

class XlsFileProcessor(SpreadsheetFileProcessor):
    file_extension = '.xls'

    def _load_sheets(self, content: bytes) -> list[Sheet]:
        wb = xlrd.open_workbook(file_contents=content)
        return [XlsSheet(sheet) for sheet in wb.sheets()]
    
class BasicPdfFileProcessor(BaseFileProcessor):
    
    def supports(self, file: File) -> bool:
        return file.name.lower().endswith('.pdf')
    
    def extract_text(self, file: File, file_quota: FileQuota) -> str:
        return process_pdf_basic(file, file_quota)

class EnhancedPdfFileProcessor(BaseFileProcessor):
    
    def supports(self, file: File) -> bool:
        return file.name.lower().endswith('.pdf')
    
    def extract_text(self, file: File, file_quota: FileQuota) -> str:
        return process_pdf_enhanced(file, file_quota)
    
class ImageFileProcessor(BaseFileProcessor):
    
    def supports(self, file: File) -> bool:
        return any(file.name.lower().endswith(ext) for ext in {'.jpg', '.jpeg', '.png'})
    
    def extract_text(self, file: File, file_quota: FileQuota) -> str:
        try:
            image_bytes = io.BytesIO(file.content)
            image = Image.open(image_bytes)
            image.verify()
        except Exception as e:
            logger.error(f"Invalid image file {file.name}: {e}")
            raise ValueError(f"Invalid image file: {file.name}")
        
        return f"Image file: {file.name}"